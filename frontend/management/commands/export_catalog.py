from django.core.management.base import BaseCommand
from frontend.models import Products, Category

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from urllib.request import urlopen
from PIL import Image as PILImage

from pathlib import Path


class Command(BaseCommand):
    help = "Export product catalogue category-wise to Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "--category",
            type=str,
            help="Category slug to export. Example: rakhi"
        )

        parser.add_argument(
            "--output",
            type=str,
            default="backup/catalogue.xlsx",
            help="Output Excel filename"
        )

    def handle(self, *args, **options):

        category_slug = options.get("category")
        output_file = options["output"]

        if category_slug:
            categories = Category.objects.filter(
                slug=category_slug
            )

            if not categories.exists():
                self.stdout.write(
                    self.style.ERROR(
                        f"Category '{category_slug}' not found."
                    )
                )
                return
        else:
            categories = Category.objects.all()

        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        product_count = 0

        for category in categories:

            products = Products.objects.filter(
                category=category
            ).select_related(
                "seller"
            ).prefetch_related(
                "media"
            )

            if not products.exists():
                continue

            # Excel sheet names cannot exceed 31 characters
            sheet_name = category.name[:31]

            # Avoid duplicate sheet names
            if sheet_name in wb.sheetnames:
                sheet_name = f"{category.id}-{sheet_name}"[:31]

            ws = wb.create_sheet(sheet_name)

            # Title
            ws["A1"] = f"{category.name} Catalogue"
            ws["A1"].font = Font(
                bold=True,
                size=18
            )

            ws.merge_cells("A1:G1")

            # Headers
            headers = [
                "Image",
                "Product",
                "Price",
                "Seller",
                "Category",
                "Description",
                "Product ID",
            ]

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(
                    row=3,
                    column=col,
                    value=header
                )

                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            row = 4

            for product in products:

                ws.cell(row=row, column=2, value=product.name)
                ws.cell(row=row, column=3, value=product.price)

                if product.seller:
                    ws.cell(
                        row=row,
                        column=4,
                        value=product.seller.store_name
                    )

                ws.cell(
                    row=row,
                    column=5,
                    value=category.name
                )

                ws.cell(
                    row=row,
                    column=6,
                    value=product.description or ""
                )

                ws.cell(
                    row=row,
                    column=7,
                    value=product.id
                )

                # Primary image
                media = product.primary_image

                if media and media.file:

                    try:
                        image_url = media.file.url

                        response = urlopen(
                            image_url,
                            timeout=15
                        )

                        image_data = response.read()

                        pil_image = PILImage.open(
                            BytesIO(image_data)
                        )

                        # Convert problematic formats
                        if pil_image.mode not in ("RGB", "RGBA"):
                            pil_image = pil_image.convert("RGB")

                        image_buffer = BytesIO()

                        pil_image.thumbnail((140, 140))

                        pil_image.save(
                            image_buffer,
                            format="PNG"
                        )

                        image_buffer.seek(0)

                        excel_image = XLImage(
                            image_buffer
                        )

                        excel_image.width = 120
                        excel_image.height = 120

                        ws.add_image(
                            excel_image,
                            f"A{row}"
                        )

                    except Exception as e:

                        self.stdout.write(
                            self.style.WARNING(
                                f"Could not download image "
                                f"for product {product.id}: {e}"
                            )
                        )

                ws.row_dimensions[row].height = 100

                row += 1
                product_count += 1

            # Column widths
            widths = {
                "A": 20,
                "B": 35,
                "C": 12,
                "D": 25,
                "E": 25,
                "F": 60,
                "G": 12,
            }

            for column, width in widths.items():
                ws.column_dimensions[column].width = width

            # Freeze header
            ws.freeze_panes = "A4"

            # Alignment
            for row_cells in ws.iter_rows(
                min_row=4,
                max_row=ws.max_row
            ):
                for cell in row_cells:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

        if not wb.sheetnames:
            self.stdout.write(
                self.style.WARNING(
                    "No products found."
                )
            )
            return

        # Create output directory
        output_path = Path(output_file)
        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        wb.save(output_file)

        self.stdout.write(
            self.style.SUCCESS(
                f"Catalogue created successfully: "
                f"{output_file}"
            )
        )

        self.stdout.write(
            self.style.SUCCESS(
                f"Products exported: {product_count}"
            )
        )